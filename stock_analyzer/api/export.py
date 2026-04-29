"""
Excel export API.

POST /api/export/excel  → download .xlsx from analysis results
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Any, Dict, List, Optional

from flask import Blueprint, Response, jsonify, request

bp = Blueprint("export", __name__, url_prefix="/api/export")


# ── Colour palette (mirrors stock_tracker.py) ────────────────────────────────

C = {
    "SBUY":  "00873D", "BUY":   "70AD47", "NEU":   "FFC000",
    "SELL":  "FF6600", "SSELL": "C00000",
    "HDR":   "1F3864", "SHDR":  "2E75B6", "SEC":   "4472C4",
    "ROW1":  "FFFFFF", "ROW2":  "DAE8F5", "BRD":   "B8CCE4",
    "WHT":   "FFFFFF", "BLK":   "1F1F1F", "GRY":   "595959",
}

CSS_TO_HEX = {
    "sig-sbuy": C["SBUY"], "sig-buy": C["BUY"], "sig-neu": C["NEU"],
    "sig-sell": C["SELL"], "sig-ssell": C["SSELL"],
}

SIG_LABEL = {
    "SBUY": "STRONG BUY", "BUY": "BUY", "NEU": "NEUTRAL",
    "SELL": "SELL",        "SSELL": "STRONG SELL",
}

SIG_TXT = {"SBUY": "WHT", "BUY": "WHT", "NEU": "BLK", "SELL": "WHT", "SSELL": "WHT"}

# Dashboard columns matching stock_tracker.py
DASH_COLS = [
    ("#",          4,   "center"),     #  1
    ("TICKER",     8,   "center"),     #  2
    ("COMPANY",   22,   "left"),       #  3
    ("SECTOR",    16,   "left"),       #  4
    ("PRICE",     11,   "center"),     #  5
    ("1 DAY",      8,   "center"),     #  6
    ("1 WEEK",     8,   "center"),     #  7
    ("1 MONTH",    9,   "center"),     #  8
    ("3 MONTHS",   9,   "center"),     #  9
    ("52W POS",   13,   "center"),     # 10
    ("RSI 14",    11,   "center"),     # 11
    ("VS MA50",    9,   "center"),     # 12
    ("VS MA200",   9,   "center"),     # 13
    ("MA CROSS",  11,   "center"),     # 14
    ("MACD",       9,   "center"),     # 15
    ("BOLL %",     9,   "center"),     # 16
    ("TRAIL P/E",  9,   "center"),     # 17
    ("FWD P/E",    9,   "center"),     # 18
    ("ADX",        7,   "center"),     # 19
    ("REGIME",     9,   "center"),     # 20
    ("ATR %",      7,   "center"),     # 21
    ("VOL RATIO",  9,   "center"),     # 22
    ("RS 1M",      8,   "center"),     # 23
    ("RS 55D",     8,   "center"),     # 24
    ("RS 3M",      8,   "center"),     # 25
    ("TREND STG",  10,  "center"),     # 26
    ("VOL REG",    9,   "center"),     # 27
    ("MKT REG",    10,  "center"),     # 28
    ("REG CHG",    12,  "center"),     # 29
    ("MOM",        7,   "center"),     # 30
    ("RISK",       7,   "center"),     # 31
    ("DIP",        7,   "center"),     # 32
    ("TECH",       8,   "center"),     # 33
    ("FUND",       8,   "center"),     # 34
    ("SENT",       8,   "center"),     # 35
    ("CONF %",     8,   "center"),     # 36
    ("ELDER",      8,   "center"),     # 37
    ("SCORE",      8,   "center"),     # 38
    ("SIGNAL",    26,   "center"),     # 39
    ("ML REGIME", 13,   "center"),     # 40
    ("ML ENTRY",   9,   "center"),     # 41
    ("ML EXIT",    9,   "center"),     # 42
    ("ML SIGNAL", 14,   "center"),     # 43
]
N_DASH_COLS = len(DASH_COLS)


# ── Matplotlib chart renderer ─────────────────────────────────────────────────

def _render_stock_chart(df, ticker: str, sym: str = "$") -> Optional[bytes]:
    """
    Render a professional multi-panel stock chart (candlestick + volume +
    RSI + MACD) using matplotlib. Returns PNG bytes or None on failure.
    """
    try:
        import matplotlib
        matplotlib.use("Agg")  # non-interactive backend
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
        import matplotlib.ticker as mticker
        from matplotlib.patches import FancyBboxPatch
        from matplotlib.lines import Line2D
        import numpy as np
        import pandas as pd
    except ImportError:
        return None

    if df is None or len(df) < 20:
        return None

    # Strip timezone for matplotlib
    df = df.copy()
    if df.index.tz is not None:
        df.index = df.index.tz_localize(None)

    # ── Theme: dark professional ─────────────────────────────────────────────
    BG       = "#1a1a2e"
    PANEL_BG = "#16213e"
    GRID     = "#1e2d4a"
    TEXT     = "#e0e0e0"
    ACCENT   = "#4fc3f7"
    UP       = "#26a69a"
    DOWN     = "#ef5350"

    plt.rcParams.update({
        "figure.facecolor":   BG,
        "axes.facecolor":     PANEL_BG,
        "axes.edgecolor":     GRID,
        "axes.labelcolor":    TEXT,
        "text.color":         TEXT,
        "xtick.color":        TEXT,
        "ytick.color":        TEXT,
        "grid.color":         GRID,
        "grid.alpha":         0.5,
        "font.family":        "sans-serif",
        "font.size":          8,
    })

    dates = df.index
    close = df["Close"].values
    opn   = df["Open"].values if "Open" in df else close
    high  = df["High"].values if "High" in df else close
    low   = df["Low"].values  if "Low"  in df else close
    vol   = df["Volume"].values if "Volume" in df else np.zeros(len(df))

    # ── Figure layout: 4 panels ──────────────────────────────────────────────
    fig = plt.figure(figsize=(16, 10), dpi=150)
    gs = fig.add_gridspec(4, 1, height_ratios=[4, 1, 1.2, 1.2],
                          hspace=0.06, left=0.06, right=0.97,
                          top=0.93, bottom=0.05)

    ax_price = fig.add_subplot(gs[0])
    ax_vol   = fig.add_subplot(gs[1], sharex=ax_price)
    ax_rsi   = fig.add_subplot(gs[2], sharex=ax_price)
    ax_macd  = fig.add_subplot(gs[3], sharex=ax_price)

    # Title
    last_px = close[-1] if len(close) else 0
    prev_px = close[-2] if len(close) > 1 else last_px
    chg = last_px - prev_px
    chg_pct = (chg / prev_px * 100) if prev_px else 0
    chg_color = UP if chg >= 0 else DOWN
    chg_arrow = "+" if chg >= 0 else ""
    fig.suptitle(
        f"  {ticker}    {sym}{last_px:,.2f}    {chg_arrow}{chg:,.2f} ({chg_arrow}{chg_pct:.2f}%)",
        fontsize=14, fontweight="bold", color=chg_color, x=0.06, ha="left",
    )

    # ══════════════════════════════════════════════════════════════════════════
    #  Panel 1: Candlestick + MAs + Bollinger Bands
    # ══════════════════════════════════════════════════════════════════════════
    x = np.arange(len(df))

    # Candlestick bodies
    up_mask   = close >= opn
    down_mask = close < opn
    body_w = 0.6

    # Up candles (green)
    ax_price.bar(x[up_mask], (close - opn)[up_mask], bottom=opn[up_mask],
                 width=body_w, color=UP, edgecolor=UP, linewidth=0.5, zorder=3)
    # Down candles (red)
    ax_price.bar(x[down_mask], (opn - close)[down_mask], bottom=close[down_mask],
                 width=body_w, color=DOWN, edgecolor=DOWN, linewidth=0.5, zorder=3)
    # Wicks
    ax_price.vlines(x[up_mask],   low[up_mask],   high[up_mask],
                    colors=UP, linewidths=0.6, zorder=2)
    ax_price.vlines(x[down_mask], low[down_mask],  high[down_mask],
                    colors=DOWN, linewidths=0.6, zorder=2)

    # Moving averages
    ma_configs = [
        ("MA20",  "#ff9800", 1.0, "MA 20"),
        ("MA50",  "#4caf50", 1.2, "MA 50"),
        ("MA200", "#f44336", 1.4, "MA 200"),
    ]
    for col, color, lw, label in ma_configs:
        if col in df.columns:
            vals = df[col].values
            mask = ~np.isnan(vals.astype(float))
            if mask.any():
                ax_price.plot(x[mask], vals[mask], color=color, linewidth=lw,
                              label=label, zorder=4)

    # Bollinger Bands (shaded)
    if "BB_Upper" in df.columns and "BB_Lower" in df.columns:
        bb_u = df["BB_Upper"].values.astype(float)
        bb_l = df["BB_Lower"].values.astype(float)
        mask = ~(np.isnan(bb_u) | np.isnan(bb_l))
        if mask.any():
            ax_price.fill_between(x[mask], bb_l[mask], bb_u[mask],
                                  alpha=0.08, color=ACCENT, zorder=1)
            ax_price.plot(x[mask], bb_u[mask], color=ACCENT, linewidth=0.5,
                          alpha=0.4, linestyle="--", zorder=2)
            ax_price.plot(x[mask], bb_l[mask], color=ACCENT, linewidth=0.5,
                          alpha=0.4, linestyle="--", label="Bollinger", zorder=2)

    ax_price.set_ylabel("Price", fontsize=9, fontweight="bold")
    ax_price.yaxis.set_major_formatter(mticker.FormatStrFormatter(f"{sym}%.2f"))
    ax_price.legend(loc="upper left", fontsize=7, framealpha=0.5,
                    facecolor=PANEL_BG, edgecolor=GRID, ncol=5)
    ax_price.grid(True, linewidth=0.3)
    ax_price.tick_params(labelbottom=False)

    # ══════════════════════════════════════════════════════════════════════════
    #  Panel 2: Volume
    # ══════════════════════════════════════════════════════════════════════════
    vol_colors = np.where(close >= opn, UP, DOWN)
    ax_vol.bar(x, vol, width=body_w, color=vol_colors, alpha=0.7, zorder=3)
    ax_vol.set_ylabel("Vol", fontsize=8, fontweight="bold")
    ax_vol.yaxis.set_major_formatter(mticker.FuncFormatter(
        lambda v, _: f"{v/1e6:.0f}M" if v >= 1e6 else f"{v/1e3:.0f}K" if v >= 1e3 else f"{v:.0f}"))
    ax_vol.grid(True, linewidth=0.3)
    ax_vol.tick_params(labelbottom=False)

    # ══════════════════════════════════════════════════════════════════════════
    #  Panel 3: RSI
    # ══════════════════════════════════════════════════════════════════════════
    if "RSI" in df.columns:
        rsi = df["RSI"].values.astype(float)
        mask = ~np.isnan(rsi)
        if mask.any():
            ax_rsi.plot(x[mask], rsi[mask], color="#ab47bc", linewidth=1.2, zorder=3)
            ax_rsi.fill_between(x[mask], 30, rsi[mask],
                                where=rsi[mask] < 30, alpha=0.15, color=UP, zorder=2)
            ax_rsi.fill_between(x[mask], 70, rsi[mask],
                                where=rsi[mask] > 70, alpha=0.15, color=DOWN, zorder=2)

    ax_rsi.axhline(70, color=DOWN, linewidth=0.8, linestyle="--", alpha=0.7)
    ax_rsi.axhline(30, color=UP,   linewidth=0.8, linestyle="--", alpha=0.7)
    ax_rsi.axhline(50, color=TEXT,  linewidth=0.4, linestyle=":", alpha=0.3)
    ax_rsi.set_ylim(0, 100)
    ax_rsi.set_ylabel("RSI", fontsize=8, fontweight="bold")
    ax_rsi.grid(True, linewidth=0.3)
    ax_rsi.tick_params(labelbottom=False)

    # Zone labels
    ax_rsi.text(len(x) - 1, 75, "OVERBOUGHT", fontsize=6, color=DOWN,
                alpha=0.6, ha="right", va="bottom")
    ax_rsi.text(len(x) - 1, 25, "OVERSOLD", fontsize=6, color=UP,
                alpha=0.6, ha="right", va="top")

    # ══════════════════════════════════════════════════════════════════════════
    #  Panel 4: MACD
    # ══════════════════════════════════════════════════════════════════════════
    if "MACD" in df.columns and "MACD_Sig" in df.columns and "MACD_Hist" in df.columns:
        macd_line = df["MACD"].values.astype(float)
        macd_sig  = df["MACD_Sig"].values.astype(float)
        macd_hist = df["MACD_Hist"].values.astype(float)

        mask = ~(np.isnan(macd_line) | np.isnan(macd_sig) | np.isnan(macd_hist))
        if mask.any():
            # Histogram bars colored by direction
            hist_colors = np.where(macd_hist >= 0,
                                   np.where(macd_hist >= np.roll(macd_hist, 1), UP, "#1b5e20"),
                                   np.where(macd_hist <= np.roll(macd_hist, 1), DOWN, "#b71c1c"))
            ax_macd.bar(x[mask], macd_hist[mask], width=body_w,
                        color=hist_colors[mask], alpha=0.7, zorder=2)
            # MACD and signal lines
            ax_macd.plot(x[mask], macd_line[mask], color="#42a5f5",
                         linewidth=1.2, label="MACD", zorder=3)
            ax_macd.plot(x[mask], macd_sig[mask], color="#ff7043",
                         linewidth=1.0, label="Signal", zorder=3)

    ax_macd.axhline(0, color=TEXT, linewidth=0.5, alpha=0.4)
    ax_macd.set_ylabel("MACD", fontsize=8, fontweight="bold")
    ax_macd.legend(loc="upper left", fontsize=7, framealpha=0.5,
                   facecolor=PANEL_BG, edgecolor=GRID)
    ax_macd.grid(True, linewidth=0.3)

    # X-axis: dates on bottom panel only
    # Show every ~20th date as label
    step = max(1, len(df) // 15)
    ax_macd.set_xticks(x[::step])
    ax_macd.set_xticklabels(
        [d.strftime("%b %d") if hasattr(d, "strftime") else str(d)
         for d in dates[::step]],
        rotation=45, fontsize=6, ha="right",
    )

    # ── Elder Impulse indicator strip at the very top ────────────────────────
    if "Elder_D" in df.columns:
        elder_map = {"green": UP, "red": DOWN, "blue": "#546e7a"}
        elder = df["Elder_D"].values
        for i in range(len(elder)):
            c = elder_map.get(str(elder[i]), "#546e7a")
            ax_price.axvspan(i - 0.4, i + 0.4, ymin=0.97, ymax=1.0,
                             color=c, alpha=0.8, zorder=5)

    # ── Save to bytes ────────────────────────────────────────────────────────
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", pad_inches=0.1)
    plt.close(fig)
    buf.seek(0)
    return buf.read()


# ── openpyxl helpers ──────────────────────────────────────────────────────────

def _build_excel(results: List[Dict], task_config: Dict,
                  raw_data: Optional[Dict] = None) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import numpy as np
    except ImportError:
        raise RuntimeError("openpyxl is required for Excel export")

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def font(color="1F1F1F", bold=False, size=10):
        return Font(color=color, bold=bold, size=size, name="Calibri")

    def align(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def border(color="B8CCE4"):
        s = Side(style="thin", color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def _col(n):
        """Convert column number to letter."""
        return get_column_letter(n)

    def _sig_hex(css: str) -> str:
        return CSS_TO_HEX.get(css) or CSS_TO_HEX.get(f"sig-{css}", C["NEU"])

    def _sig_label(signal: str) -> str:
        return SIG_LABEL.get(signal, signal or "NEUTRAL")

    def _sig_txt(css: str) -> str:
        return "FFFFFF" if css in ("sig-sbuy", "sig-buy", "sig-sell", "sig-ssell",
                                   "sbuy", "buy", "sell", "ssell") else "1F1F1F"

    def _sent_hex(score) -> str:
        if score is None:
            return C["GRY"]
        if score >= 0.5:   return C["SBUY"]
        if score >= 0.2:   return C["BUY"]
        if score >= -0.2:  return C["NEU"]
        if score >= -0.5:  return C["SELL"]
        return C["SSELL"]

    def _score_color(score) -> str:
        """Map a [-1, +1] score to a background hex."""
        if score is None:  return C["GRY"]
        if score >= 0.5:   return C["SBUY"]
        if score >= 0.2:   return C["BUY"]
        if score >= -0.2:  return C["NEU"]
        if score >= -0.5:  return C["SELL"]
        return C["SSELL"]

    def _sent_label(signal: Optional[str]) -> str:
        if not signal:
            return "N/A"
        m = {"BULLISH": "BULLISH", "NEUTRAL": "NEUTRAL", "BEARISH": "BEARISH",
             "STRONG_BULLISH": "STR. BULLISH", "STRONG_BEARISH": "STR. BEARISH"}
        return m.get(signal.upper(), signal)

    def _pct(v):
        return None if v is None else round(v, 2)

    def _r(v, d=2):
        return None if v is None else round(v, d)

    def _safe(v):
        return v if v is not None else 0

    # ── Colour functions (mirror stock_tracker.py) ────────────────────────────

    def _pct_color(v):
        if v is None: return None
        if v >  5:  return C["SBUY"]
        if v >  0:  return C["BUY"]
        if v < -5:  return C["SSELL"]
        if v <  0:  return C["SELL"]
        return C["NEU"]

    def _rsi_color(v):
        if v is None: return None
        if v < 30:  return C["SBUY"]
        if v < 45:  return C["BUY"]
        if v > 70:  return C["SSELL"]
        if v > 55:  return C["SELL"]
        return C["NEU"]

    def _pe_color(v):
        if v is None or v <= 0: return None
        if v < 10:  return C["SBUY"]
        if v < 20:  return C["BUY"]
        if v < 30:  return C["NEU"]
        if v < 50:  return C["SELL"]
        return C["SSELL"]

    def _margin_color(v):
        if v is None: return None
        if v >= 25:  return C["SBUY"]
        if v >= 10:  return C["BUY"]
        if v >=  0:  return C["NEU"]
        if v >= -10: return C["SELL"]
        return C["SSELL"]

    def _growth_color(v):
        if v is None: return None
        if v >= 20:  return C["SBUY"]
        if v >=  5:  return C["BUY"]
        if v >=  0:  return C["NEU"]
        if v >= -10: return C["SELL"]
        return C["SSELL"]

    def _de_color(v):
        if v is None: return None
        if v < 30:   return C["SBUY"]
        if v < 100:  return C["BUY"]
        if v < 200:  return C["NEU"]
        if v < 400:  return C["SELL"]
        return C["SSELL"]

    def _cr_color(v):
        if v is None: return None
        if v >= 3.0: return C["SBUY"]
        if v >= 1.5: return C["BUY"]
        if v >= 1.0: return C["NEU"]
        return C["SSELL"]

    def _bar10(pct):
        if pct is None: return "N/A"
        filled = max(0, min(10, round(pct / 10)))
        return "█" * filled + "░" * (10 - filled) + f"  {pct:.0f}%"

    def _pct_str(v, d=2):
        if v is None: return "N/A"
        arrow = "▲ " if v >= 0 else "▼ "
        return f"{arrow}{v:+.{d}f}%"

    def _pe_str(v):
        if v is None: return "N/A"
        if v <= 0: return "Loss"
        return f"{v:.1f}x"

    def _fmt_cap(v):
        if v is None: return "N/A"
        if v >= 1e12: return f"${v/1e12:.2f}T"
        if v >= 1e9:  return f"${v/1e9:.1f}B"
        if v >= 1e6:  return f"${v/1e6:.0f}M"
        return f"${v:,.0f}"

    def _rec_str(v):
        if v is None: return "N/A"
        labels = {1: "Strong Buy", 2: "Buy", 3: "Hold", 4: "Underperform", 5: "Sell"}
        return labels.get(round(v), f"{v:.1f}")

    # Helper for remaining sheets (Sentiment, Stock sheets)
    def sc(ws, row, col, value=None, bg=None, fg="1F1F1F",
           bold=False, sz=10, h="center", v="center", wrap=False, nfmt=None, italic=False):
        cell = ws.cell(row=row, column=col)
        if value is not None:
            cell.value = value
        if bg:
            cell.fill = fill(bg)
        cell.font = font(fg, bold, sz)
        cell.alignment = align(h, v, wrap)
        cell.border = border()
        if nfmt:
            cell.number_format = nfmt
        return cell

    wb = Workbook()

    # ── Dashboard sheet (matching stock_tracker.py format) ─────────────────────
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M")

    # Row 1: Title banner
    ws.merge_cells(f"A1:{_col(N_DASH_COLS)}1")
    t = ws["A1"]
    t.value = "  STOCK ANALYSIS DASHBOARD  —  Technical Analysis Overview"
    t.fill = fill(C["HDR"])
    t.font = font(C["WHT"], bold=True, size=18)
    t.alignment = align("left", "center")
    ws.row_dimensions[1].height = 42

    # Row 2: Subtitle / timestamp
    ws.merge_cells(f"A2:{_col(N_DASH_COLS)}2")
    s = ws["A2"]
    s.value = (f"  Last updated: {now_str}  │  "
               f"Source: Yahoo Finance  │  History: 1 year  │  "
               f"Scores range from −9 (extreme bearish) to +9 (extreme bullish)")
    s.fill = fill(C["SHDR"])
    s.font = font(C["WHT"], size=9)
    s.alignment = align("left", "center")
    ws.row_dimensions[2].height = 20

    # Row 3: Signal legend
    ws.row_dimensions[3].height = 24
    ws.merge_cells("A3:B3")
    lbl = ws["A3"]
    lbl.value = "SIGNAL LEGEND →"
    lbl.fill = fill(C["HDR"])
    lbl.font = font(C["WHT"], bold=True, size=9)
    lbl.alignment = align("center", "center")

    leg = [
        ("STRONG BUY",  "SBUY"),
        ("BUY",         "BUY"),
        ("NEUTRAL",     "NEU"),
        ("SELL",        "SELL"),
        ("STRONG SELL", "SSELL"),
    ]
    cs = 3
    span = (N_DASH_COLS - 2) // len(leg)
    for text, key in leg:
        e = cs + span - 1
        if e > N_DASH_COLS:
            e = N_DASH_COLS
        ws.merge_cells(f"{_col(cs)}3:{_col(e)}3")
        cell = ws.cell(row=3, column=cs)
        cell.value = text
        cell.fill = fill(C[key])
        cell.font = font(C[SIG_TXT[key]], bold=True, size=10)
        cell.alignment = align("center", "center")
        cs += span
    # Fill remainder
    for ci in range(cs, N_DASH_COLS + 1):
        ws.cell(row=3, column=ci).fill = fill(C["HDR"])

    # Row 4: thin separator
    ws.row_dimensions[4].height = 6
    for ci in range(1, N_DASH_COLS + 1):
        ws.cell(row=4, column=ci).fill = fill(C["HDR"])

    # Row 5: Column headers
    ws.row_dimensions[5].height = 30
    for ci, (hdr, w, ha) in enumerate(DASH_COLS, start=1):
        cell = ws.cell(row=5, column=ci)
        cell.value = hdr
        cell.fill = fill(C["SHDR"])
        cell.font = font(C["WHT"], bold=True, size=10)
        cell.alignment = align(ha, "center", wrap=True)
        cell.border = border(C["WHT"])
        ws.column_dimensions[_col(ci)].width = w

    # Rows 6+: Data rows
    for ri, r in enumerate(results):
        row = 6 + ri
        ws.row_dimensions[row].height = 22
        even = ri % 2 == 0
        bg = C["ROW2"] if even else C["ROW1"]

        sym = "€" if r.get("currency") == "EUR" else "$"
        px = r.get("price")
        ma50_v = r.get("ma50")
        ma200_v = r.get("ma200")

        def pct_vs(px, ma):
            if px and ma:
                return round((px - ma) / ma * 100, 1)
            return None

        vs50 = pct_vs(px, ma50_v)
        vs200 = pct_vs(px, ma200_v)

        # Helper to put cell data
        def put(ci, value, bg_=None, fg_=C["BLK"], bold=False, nfmt=None, ha="center"):
            cell = ws.cell(row=row, column=ci)
            cell.value = value
            cell.fill = fill(bg_ or bg)
            cell.font = font(fg_, bold=bold, size=10)
            cell.alignment = align(ha, "center")
            cell.border = border()
            if nfmt:
                cell.number_format = nfmt

        # Column 1: #
        put(1, ri + 1)

        # Column 2: TICKER
        put(2, r.get("ticker"), bold=True)

        # Column 3: COMPANY
        put(3, r.get("name"), ha="left")

        # Column 4: SECTOR
        put(4, r.get("sector"), fg_=C["GRY"], ha="left")

        # Column 5: PRICE
        if px:
            put(5, px, bold=True, nfmt=f'{sym}#,##0.00')
        else:
            put(5, "N/A")

        # Columns 6-9: Returns (1D, 1W, 1M, 3M) — background colour by magnitude
        for ci, ret_key in enumerate(["ret_1d", "ret_1w", "ret_1m", "ret_3m"], 6):
            v = r.get(ret_key)
            pb = _pct_color(v)
            pf = C["WHT"] if pb and pb != C["NEU"] else C["BLK"]
            put(ci, _pct_str(v) if v is not None else "N/A", bg_=pb, fg_=pf, bold=bool(pb))

        # Column 10: 52W POS — unicode bar
        w52 = r.get("w52_pct")
        w52_bg = C["BUY"] if (w52 or 0) >= 50 else C["SELL"]
        put(10, _bar10(w52), bg_=w52_bg, fg_=C["WHT"])

        # Column 11: RSI 14 — background + OVERSOLD/OVERBOUGHT label
        rsi = r.get("rsi")
        rb = _rsi_color(rsi)
        rf = C["WHT"] if rb and rb != C["NEU"] else C["BLK"]
        rsi_lbl = "N/A"
        if rsi is not None:
            rsi_lbl = f"{rsi:.1f}"
            if   rsi < 30: rsi_lbl += "  OVERSOLD"
            elif rsi > 70: rsi_lbl += "  OVERBOUGHT"
        put(11, rsi_lbl, bg_=rb, fg_=rf, bold=True)

        # Column 12: VS MA50 — background + arrow text
        if px and ma50_v:
            pb = C["BUY"] if vs50 >= 0 else C["SELL"]
            put(12, f"{'▲' if vs50 >= 0 else '▼'} {vs50:+.1f}%", bg_=pb, fg_=C["WHT"], bold=True)
        else:
            put(12, "N/A")

        # Column 13: VS MA200 — background + arrow text
        if px and ma200_v:
            pb = C["BUY"] if vs200 >= 0 else C["SSELL"]
            put(13, f"{'▲' if vs200 >= 0 else '▼'} {vs200:+.1f}%", bg_=pb, fg_=C["WHT"], bold=True)
        else:
            put(13, "N/A")

        # Column 14: MA CROSS
        cross = r.get("ma_cross")
        if cross == "golden":
            put(14, "✓ GOLDEN", bg_=C["BUY"], fg_=C["WHT"], bold=True)
        elif cross == "death":
            put(14, "✗ DEATH", bg_=C["SSELL"], fg_=C["WHT"], bold=True)
        else:
            put(14, "N/A")

        # Column 15: MACD — with arrows
        macd_b = r.get("macd_bull")
        if macd_b is True:
            put(15, "▲ BULL", bg_=C["BUY"], fg_=C["WHT"], bold=True)
        elif macd_b is False:
            put(15, "▼ BEAR", bg_=C["SELL"], fg_=C["WHT"], bold=True)
        else:
            put(15, "N/A")

        # Column 16: BOLL % — background colour + text
        bbp = r.get("bb_pct")
        if bbp is not None:
            bb_bg = C["SBUY"] if bbp < 15 else C["SSELL"] if bbp > 85 else C["NEU"]
            bb_fg = C["WHT"] if bbp < 15 or bbp > 85 else C["BLK"]
            put(16, f"{bbp:.1f}%", bg_=bb_bg, fg_=bb_fg, bold=True)
        else:
            put(16, "N/A")

        # Column 17: TRAIL P/E — background colour + "x" suffix
        pe = r.get("pe_trail")
        peb = _pe_color(pe)
        pef_c = C["WHT"] if peb and peb != C["NEU"] else C["BLK"]
        put(17, _pe_str(pe), bg_=peb, fg_=pef_c, bold=bool(peb))

        # Column 18: FWD P/E
        pef_v = r.get("pe_fwd")
        fpeb = _pe_color(pef_v)
        fpef_c = C["WHT"] if fpeb and fpeb != C["NEU"] else C["BLK"]
        put(18, _pe_str(pef_v), bg_=fpeb, fg_=fpef_c, bold=bool(fpeb))

        # Column 19: ADX
        adx_v = r.get("adx")
        if adx_v is not None:
            adx_bg = C["SBUY"] if adx_v > 25 else C["NEU"] if adx_v >= 20 else C["GRY"]
            adx_fg = C["WHT"] if adx_v > 25 else C["BLK"]
            put(19, f"{adx_v:.1f}", bg_=adx_bg, fg_=adx_fg, bold=adx_v > 25)
        else:
            put(19, "N/A", fg_=C["GRY"])

        # Column 20: REGIME
        regime_v = r.get("regime", "NEUTRAL")
        regime_map = {"TREND": (C["SBUY"], "TREND"), "MEAN_REVERSION": (C["SHDR"], "MEAN REV"), "NEUTRAL": (C["NEU"], "NEUTRAL")}
        reg_bg, reg_lbl = regime_map.get(regime_v, (C["GRY"], regime_v))
        reg_fg = C["WHT"] if regime_v == "TREND" else C["BLK"] if regime_v == "NEUTRAL" else C["WHT"]
        put(20, reg_lbl, bg_=reg_bg, fg_=reg_fg, bold=True)

        # Column 21: ATR %
        atr_pct_v = r.get("atr_pct")
        if atr_pct_v is not None:
            atr_bg = C["SSELL"] if atr_pct_v > 5 else C["SELL"] if atr_pct_v > 3 else C["NEU"] if atr_pct_v > 1.5 else C["BUY"]
            atr_fg = C["WHT"] if atr_pct_v > 3 else C["BLK"]
            put(21, f"{atr_pct_v:.1f}%", bg_=atr_bg, fg_=atr_fg, bold=atr_pct_v > 3)
        else:
            put(21, "N/A", fg_=C["GRY"])

        # Column 22: VOL RATIO
        vr_v = r.get("vol_ratio")
        if vr_v is not None:
            vr_bg = C["SBUY"] if vr_v > 2.0 else C["BUY"] if vr_v > 1.5 else C["SELL"] if vr_v < 0.5 else None
            vr_fg = C["WHT"] if vr_bg and vr_bg != C["NEU"] else C["BLK"]
            put(22, f"{vr_v:.2f}x", bg_=vr_bg, fg_=vr_fg, bold=vr_v > 1.5 or vr_v < 0.5)
        else:
            put(22, "N/A", fg_=C["GRY"])

        # Column 23: RS 1M (relative strength vs SPY, 21-day)
        rs_v = r.get("rs_1m")
        if rs_v is not None:
            rs_bg = C["SBUY"] if rs_v > 10 else C["BUY"] if rs_v > 2 else C["SSELL"] if rs_v < -10 else C["SELL"] if rs_v < -2 else None
            rs_fg = C["WHT"] if rs_bg else C["BLK"]
            put(23, f"{rs_v:+.1f}%", bg_=rs_bg, fg_=rs_fg, bold=abs(rs_v) > 5)
        else:
            put(23, "N/A", fg_=C["GRY"])

        # Column 24: RS 55D (relative strength vs SPY, 55-day / IBD-inspired)
        rs55_v = r.get("rs_55d")
        if rs55_v is not None:
            rs55_bg = C["SBUY"] if rs55_v > 10 else C["BUY"] if rs55_v > 2 else C["SSELL"] if rs55_v < -10 else C["SELL"] if rs55_v < -2 else None
            rs55_fg = C["WHT"] if rs55_bg else C["BLK"]
            put(24, f"{rs55_v:+.1f}%", bg_=rs55_bg, fg_=rs55_fg, bold=abs(rs55_v) > 5)
        else:
            put(24, "N/A", fg_=C["GRY"])

        # Column 25: RS 3M (relative strength vs SPY, 63-day)
        rs3_v = r.get("rs_3m")
        if rs3_v is not None:
            rs3_bg = C["SBUY"] if rs3_v > 10 else C["BUY"] if rs3_v > 2 else C["SSELL"] if rs3_v < -10 else C["SELL"] if rs3_v < -2 else None
            rs3_fg = C["WHT"] if rs3_bg else C["BLK"]
            put(25, f"{rs3_v:+.1f}%", bg_=rs3_bg, fg_=rs3_fg, bold=abs(rs3_v) > 5)
        else:
            put(25, "N/A", fg_=C["GRY"])

        # Column 26: TREND STAGE
        ts_v = r.get("trend_stage")
        if ts_v:
            ts_map = {"EARLY": (C["BUY"], C["WHT"]), "HEALTHY": (C["SBUY"], C["WHT"]),
                       "EXTENDED": (C["NEU"], C["BLK"]), "OVEREXTENDED": (C["SELL"], C["WHT"]),
                       "PARABOLIC": (C["SSELL"], C["WHT"])}
            ts_bg, ts_fg = ts_map.get(ts_v, (C["GRY"], C["BLK"]))
            put(26, ts_v, bg_=ts_bg, fg_=ts_fg, bold=True)
        else:
            put(26, "N/A", fg_=C["GRY"])

        # Column 27: VOL REGIME
        vr_regime = r.get("vol_regime")
        if vr_regime:
            vr_map = {"LOW": (C["BUY"], C["WHT"]), "NORMAL": (C["NEU"], C["BLK"]),
                       "HIGH": (C["SELL"], C["WHT"]), "EXTREME": (C["SSELL"], C["WHT"])}
            vr_bg2, vr_fg2 = vr_map.get(vr_regime, (C["GRY"], C["BLK"]))
            put(27, vr_regime, bg_=vr_bg2, fg_=vr_fg2, bold=True)
        else:
            put(27, "N/A", fg_=C["GRY"])

        # Column 28: MKT REGIME
        mr_v = r.get("mkt_regime")
        if mr_v:
            mr_map = {"BULLISH": (C["SBUY"], C["WHT"]), "BEARISH": (C["SSELL"], C["WHT"]),
                       "TRANSITION": (C["NEU"], C["BLK"])}
            mr_bg, mr_fg = mr_map.get(mr_v, (C["GRY"], C["BLK"]))
            put(28, mr_v, bg_=mr_bg, fg_=mr_fg, bold=True)
        else:
            put(28, "N/A", fg_=C["GRY"])

        # Column 29: REGIME CHG
        rc_v = r.get("regime_chg")
        if rc_v:
            rc_map = {"BEARISH REVERSAL": (C["SSELL"], C["WHT"]),
                       "BULLISH REVERSAL": (C["SBUY"], C["WHT"]),
                       "WEAKENING": (C["SELL"], C["WHT"]),
                       "POTENTIAL BOTTOM": (C["BUY"], C["WHT"]),
                       "BULLISH CONFIRMATION": (C["BUY"], C["WHT"]),
                       "BEARISH CONFIRMATION": (C["SELL"], C["WHT"])}
            rc_bg, rc_fg = rc_map.get(rc_v, (C["GRY"], C["BLK"]))
            put(29, rc_v, bg_=rc_bg, fg_=rc_fg, bold=True)
        else:
            put(29, "—", fg_=C["GRY"])

        # Column 30: MOM (momentum score 0-1)
        mom_v = r.get("momentum_score")
        if mom_v is not None:
            mom_bg = C["SBUY"] if mom_v >= 0.65 else C["BUY"] if mom_v >= 0.40 else C["NEU"] if mom_v >= 0.20 else C["GRY"]
            mom_fg = C["WHT"] if mom_v >= 0.40 else C["BLK"]
            put(30, f"{mom_v:.2f}", bg_=mom_bg, fg_=mom_fg, bold=mom_v >= 0.65)
        else:
            put(30, "N/A", fg_=C["GRY"])

        # Column 31: RISK (risk interaction score)
        risk_v = r.get("risk_score")
        if risk_v is not None:
            risk_bg = C["SSELL"] if risk_v >= 1.0 else C["SELL"] if risk_v >= 0.5 else C["NEU"] if risk_v >= 0.15 else C["BUY"]
            risk_fg = C["WHT"] if risk_v >= 0.5 else C["BLK"]
            put(31, f"{risk_v:.2f}", bg_=risk_bg, fg_=risk_fg, bold=risk_v >= 0.5)
        else:
            put(31, "N/A", fg_=C["GRY"])

        # Column 32: DIP (dip quality score 0-1)
        dip_v = r.get("dip_score")
        if dip_v is not None and dip_v > 0:
            dip_bg = C["SBUY"] if dip_v >= 0.65 else C["BUY"] if dip_v >= 0.35 else C["NEU"]
            dip_fg = C["WHT"] if dip_v >= 0.35 else C["BLK"]
            put(32, f"{dip_v:.2f}", bg_=dip_bg, fg_=dip_fg, bold=dip_v >= 0.55)
        else:
            put(32, "—", fg_=C["GRY"])

        # Column 33: TECH score [-1, +1]
        tech_s = r.get("tech_score")
        if tech_s is not None:
            tb = _score_color(tech_s)
            tf = C["WHT"] if abs(tech_s) >= 0.2 else C["BLK"]
            put(33, f"{tech_s:+.2f}", bg_=tb, fg_=tf, bold=True)
        else:
            put(33, "N/A", fg_=C["GRY"])

        # Column 34: FUND score [-1, +1]
        fund_s = r.get("fund_score")
        if fund_s is not None:
            fb = _score_color(fund_s)
            ff = C["WHT"] if abs(fund_s) >= 0.2 else C["BLK"]
            put(34, f"{fund_s:+.2f}", bg_=fb, fg_=ff, bold=True)
        else:
            put(34, "N/A", fg_=C["GRY"])

        # Column 35: SENT score [-1, +1]
        ss = r.get("sent_score")
        if ss is not None:
            sh = _sent_hex(ss)
            sf = C["WHT"] if abs(ss) >= 0.2 else C["BLK"]
            put(35, f"{ss:+.2f}", bg_=sh, fg_=sf, bold=True)
        else:
            put(35, "N/A", fg_=C["GRY"])

        # Column 36: CONF % — adjusted confidence
        conf = r.get("adj_confidence") or r.get("confidence")
        if conf is not None:
            cb = C["SBUY"] if conf >= 70 else C["BUY"] if conf >= 50 else C["NEU"] if conf >= 30 else C["SELL"]
            cf = C["WHT"] if conf >= 50 or conf < 30 else C["BLK"]
            put(36, f"{conf:.0f}%", bg_=cb, fg_=cf, bold=True)
        else:
            put(36, "N/A", fg_=C["GRY"])

        # Column 37: ELDER — weekly Elder Impulse (G/B/R)
        elder_w = r.get("elder_w")
        if elder_w:
            e_map = {"green": (C["SBUY"], "GREEN"), "red": (C["SSELL"], "RED"), "blue": (C["SHDR"], "BLUE")}
            e_bg, e_lbl = e_map.get(elder_w, (C["GRY"], elder_w.upper()))
            put(37, e_lbl, bg_=e_bg, fg_=C["WHT"], bold=True)
        else:
            put(37, "N/A", fg_=C["GRY"])

        # Column 38: SCORE — overall weighted score [-1, +1]
        ov_score = r.get("overall_score") or 0
        ov_css = r.get("signal_css", "neu")
        ov_bg = _sig_hex(ov_css)
        ov_fg = _sig_txt(ov_css)
        put(38, f"{ov_score:+.2f}", bg_=ov_bg, fg_=ov_fg, bold=True)

        # Column 39: SIGNAL — contextual label (v4.1 decision tree)
        css = r.get("signal_css", "sig-neu")
        ctx_sig = r.get("ctx_signal") or ""
        if "AVOID" in ctx_sig:
            sig_bg = _sig_hex("ssell")
            sig_fg = _sig_txt("ssell")
        elif ctx_sig.startswith("HOLD"):
            sig_bg = _sig_hex("neu")
            sig_fg = _sig_txt("neu")
        elif "DIP" in ctx_sig or "BOTTOM" in ctx_sig or "REVERSAL" in ctx_sig:
            sig_bg = _sig_hex("buy")
            sig_fg = _sig_txt("buy")
        else:
            sig_bg = _sig_hex(css)
            sig_fg = _sig_txt(css)
        sig_lbl = ctx_sig or _sig_label(r.get("signal") or "")
        put(39, sig_lbl, bg_=sig_bg, fg_=sig_fg, bold=True)

        # Column 40: ML REGIME
        ml_regime = r.get("ml_regime")
        if ml_regime:
            ml_reg_map = {
                "TREND_UP":       (C["SBUY"],  C["WHT"]),
                "TREND_DOWN":     (C["SSELL"], C["WHT"]),
                "REVERSAL_UP":    (C["BUY"],   C["WHT"]),
                "REVERSAL_DOWN":  (C["SELL"],  C["WHT"]),
                "RANGE":          (C["NEU"],   C["BLK"]),
            }
            ml_rb, ml_rf = ml_reg_map.get(ml_regime, (C["GRY"], C["BLK"]))
            put(40, ml_regime.replace("_", " "), bg_=ml_rb, fg_=ml_rf, bold=True)
        else:
            put(40, "—", fg_=C["GRY"])

        # Column 41: ML ENTRY score
        ml_entry = r.get("ml_entry_score")
        if ml_entry is not None:
            ml_eb = C["SBUY"] if ml_entry >= 0.75 else C["BUY"] if ml_entry >= 0.6 else C["NEU"] if ml_entry >= 0.4 else C["SELL"]
            ml_ef = C["WHT"] if ml_entry >= 0.6 or ml_entry < 0.4 else C["BLK"]
            put(41, f"{ml_entry:.0%}", bg_=ml_eb, fg_=ml_ef, bold=ml_entry >= 0.6)
        else:
            put(41, "—", fg_=C["GRY"])

        # Column 42: ML EXIT score
        ml_exit = r.get("ml_exit_score")
        if ml_exit is not None:
            ml_xb = C["SSELL"] if ml_exit >= 0.75 else C["SELL"] if ml_exit >= 0.6 else C["NEU"] if ml_exit >= 0.4 else C["BUY"]
            ml_xf = C["WHT"] if ml_exit >= 0.6 or ml_exit < 0.4 else C["BLK"]
            put(42, f"{ml_exit:.0%}", bg_=ml_xb, fg_=ml_xf, bold=ml_exit >= 0.6)
        else:
            put(42, "—", fg_=C["GRY"])

        # Column 43: ML SIGNAL (BUY / HOLD / SELL / REDUCE)
        ml_sig = r.get("ml_signal")
        if ml_sig:
            ml_sig_map = {
                "BUY":    (C["SBUY"],  C["WHT"]),
                "REDUCE": (C["SELL"],  C["WHT"]),
                "SELL":   (C["SSELL"], C["WHT"]),
                "HOLD":   (C["NEU"],   C["BLK"]),
            }
            ml_sb, ml_sf = ml_sig_map.get(ml_sig.upper(), (C["GRY"], C["BLK"]))
            put(43, ml_sig.upper(), bg_=ml_sb, fg_=ml_sf, bold=True)
        else:
            put(43, "—", fg_=C["GRY"])

    # ── Footer row ────────────────────────────────────────────────────────────
    fr = 6 + len(results)
    ws.row_dimensions[fr].height = 14
    ws.merge_cells(f"A{fr}:{_col(N_DASH_COLS)}{fr}")
    fc = ws.cell(row=fr, column=1)
    fc.value = (
        "RSI < 30 = Oversold (possible buy opportunity)  │  RSI > 70 = Overbought (possible sell)  │  "
        "Golden Cross = MA50 above MA200 (bullish)  │  Death Cross = MA50 below MA200 (bearish)  │  "
        "BB% < 15 = near lower band (oversold)  │  BB% > 85 = near upper band (overbought)"
    )
    fc.fill      = fill(C["HDR"])
    fc.font      = font("AAAAAA", size=8)
    fc.alignment = align("center", "center")

    # Check if any results have sentiment data
    has_sent = any(r.get("sent_score") is not None for r in results)

    # ── Fundamentals sheet ────────────────────────────────────────────────────
    ws_f = wb.create_sheet("Fundamentals")
    ws_f.sheet_view.showGridLines = False
    ws_f.freeze_panes = "C5"

    # Fundamentals sections (simplified from stock_tracker.py)
    FUND_LEFT = [("#", 4), ("TICKER", 9), ("COMPANY", 22)]
    FUND_COLS = [
        ("MARKET & RISK", "1F3864", [
            ("MKT CAP", 12, "mkt_cap"), ("BETA", 7, "beta"),
            ("DIV YIELD", 9, "div_yield"), ("SHORT FLOAT", 9, "short_float"),
        ]),
        ("VALUATION", "7B3F00", [
            ("TRAIL P/E", 9, "pe_trail"), ("FWD P/E", 9, "pe_fwd"),
            ("PEG", 7, "peg"), ("P/BOOK", 8, "pb"),
        ]),
        ("PROFITABILITY", "1E5631", [
            ("NET MGN", 9, "net_mgn"), ("ROE", 8, "roe"), ("ROA", 8, "roa"),
        ]),
        ("GROWTH", "4A235A", [
            ("REV GROWTH", 9, "rev_growth"), ("EPS GROWTH", 9, "eps_growth"),
        ]),
        ("FINANCIAL HEALTH", "7B3F00", [
            ("D/E", 9, "debt_eq"), ("CURR RATIO", 9, "curr_ratio"),
        ]),
        ("ANALYST", "154360", [
            ("TARGET", 9, "target_px"), ("RATING", 10, "rec_mean"), ("# ANALYSTS", 9, "n_analysts"),
        ]),
    ]

    N_FUND_LEFT = len(FUND_LEFT)
    all_fund_cols = []
    for sec_label, sec_color, cols in FUND_COLS:
        for ch, w, dk in cols:
            all_fund_cols.append((sec_label, sec_color, ch, w, dk))
    N_FUND_TOTAL = N_FUND_LEFT + len(all_fund_cols)

    # Title
    ws_f.merge_cells(f"A1:{_col(N_FUND_TOTAL)}1")
    t = ws_f["A1"]
    t.value = "  FUNDAMENTALS OVERVIEW  —  Valuation · Profitability · Growth · Health · Analyst"
    t.fill = fill(C["HDR"])
    t.font = font(C["WHT"], bold=True, size=15)
    t.alignment = align("left", "center")
    ws_f.row_dimensions[1].height = 36

    # Subtitle
    ws_f.merge_cells(f"A2:{_col(N_FUND_TOTAL)}2")
    s = ws_f["A2"]
    s.value = "  Fundamentals data from Yahoo Finance"
    s.fill = fill(C["SHDR"])
    s.font = font(C["WHT"], size=9)
    s.alignment = align("left", "center")
    ws_f.row_dimensions[2].height = 20

    # Section headers (row 3)
    ws_f.row_dimensions[3].height = 20
    for ci, (hdr, w) in enumerate(FUND_LEFT, 1):
        cell = ws_f.cell(row=3, column=ci)
        cell.value = hdr
        cell.fill = fill(C["HDR"])
        cell.font = font(C["WHT"], bold=True, size=10)
        cell.alignment = align("center", "center")
        ws_f.column_dimensions[_col(ci)].width = w

    ci = N_FUND_LEFT + 1
    cur_sec = None
    sec_start = ci
    for sec_label, sec_color, ch, w, dk in all_fund_cols:
        if sec_label != cur_sec:
            if cur_sec is not None:
                ws_f.merge_cells(f"{_col(sec_start)}3:{_col(ci-1)}3")
            cur_sec = sec_label
            sec_start = ci
            cell = ws_f.cell(row=3, column=ci)
            cell.value = sec_label
            cell.fill = fill(sec_color)
            cell.font = font(C["WHT"], bold=True, size=9)
            cell.alignment = align("center", "center")
        ws_f.column_dimensions[_col(ci)].width = w
        ci += 1
    if cur_sec is not None:
        ws_f.merge_cells(f"{_col(sec_start)}3:{_col(ci-1)}3")

    # Column headers (row 4)
    ws_f.row_dimensions[4].height = 24
    for ci in range(1, N_FUND_LEFT + 1):
        cell = ws_f.cell(row=4, column=ci)
        cell.fill = fill(C["HDR"])
        cell.border = border(C["WHT"])

    for ci, (sec_label, sec_color, ch, w, dk) in enumerate(all_fund_cols, N_FUND_LEFT + 1):
        cell = ws_f.cell(row=4, column=ci)
        cell.value = ch
        cell.fill = fill(sec_color)
        cell.font = font(C["WHT"], bold=True, size=9)
        cell.alignment = align("center", "center", wrap=True)
        cell.border = border(C["WHT"])

    # Data rows
    for ri, r in enumerate(results):
        row = 5 + ri
        even = ri % 2 == 0
        bg = C["ROW2"] if even else C["ROW1"]
        ws_f.row_dimensions[row].height = 20

        # Fixed columns
        ws_f.cell(row=row, column=1).value = ri + 1
        ws_f.cell(row=row, column=2).value = r.get("ticker")
        ws_f.cell(row=row, column=2).font = font(bold=True)
        ws_f.cell(row=row, column=3).value = r.get("name")
        ws_f.cell(row=row, column=3).alignment = align("left", "center")
        for ci in range(1, 4):
            ws_f.cell(row=row, column=ci).fill = fill(bg)
            ws_f.cell(row=row, column=ci).border = border()

        # Metric columns
        for ci, (sec_label, sec_color, ch, w, dk) in enumerate(all_fund_cols, N_FUND_LEFT + 1):
            val = r.get(dk)
            cell = ws_f.cell(row=row, column=ci)
            cell.fill = fill(bg)
            cell.border = border()
            cell.alignment = align("center", "center")

            _fund_color_fns = {
                "pe_trail": _pe_color, "pe_fwd": _pe_color,
                "net_mgn": _margin_color, "roe": _margin_color, "roa": _margin_color,
                "rev_growth": _growth_color, "eps_growth": _growth_color,
                "debt_eq": _de_color, "curr_ratio": _cr_color,
            }
            cfn = _fund_color_fns.get(dk)
            cb = cfn(val) if cfn and val is not None else None
            cf = C["WHT"] if cb and cb != C["NEU"] else C["BLK"]
            if cb:
                cell.fill = fill(cb)
                cell.font = font(cf, bold=True, size=10)
            if val is None:
                cell.value = "N/A"
                cell.font = font(C["GRY"])
            elif dk == "mkt_cap":
                cell.value = _fmt_cap(val)
            elif "pe" in dk or dk == "peg" or dk == "pb":
                cell.value = _pe_str(val) if "pe" in dk else f"{val:.1f}x" if val > 0 else "N/A"
            elif "mgn" in dk or "growth" in dk or dk in ("roe", "roa", "div_yield", "short_float"):
                cell.value = f"{val:+.1f}%" if val != 0 else "0.0%"
            elif dk == "debt_eq":
                cell.value = f"{val:.1f}"
            elif dk == "curr_ratio":
                cell.value = f"{val:.2f}"
            elif dk == "rec_mean":
                cell.value = _rec_str(val)
            elif dk == "n_analysts":
                cell.value = str(int(val))
            else:
                cell.value = val
                cell.number_format = '0.00'

    # ── Sentiment sheet ───────────────────────────────────────────────────────
    if has_sent:
        sent_results = [r for r in results if r.get("sent_score") is not None]
        if sent_results:
            ws2 = wb.create_sheet("Sentiment")
            ws2.freeze_panes = "A4"

            ws2.row_dimensions[1].height = 30
            ws2.merge_cells("A1:I1")
            sc(ws2, 1, 1, "SENTIMENT ANALYSIS", bg=C["HDR"], fg=C["WHT"], bold=True, sz=14)

            ws2.row_dimensions[2].height = 16
            ws2.merge_cells("A2:I2")
            sc(ws2, 2, 1, f"Generated  {now_str}", bg=C["SHDR"], fg=C["WHT"], sz=9, italic=True)

            SENT_COLS = [
                ("#",          4),  ("TICKER",    9),  ("COMPANY",  20),
                ("SCORE",      9),  ("SIGNAL",   14),  ("ARTICLES",  8),
                ("MOMENTUM",   10), ("TREND",    10),  ("DISPERSION", 10),
            ]
            ws2.row_dimensions[3].height = 24
            for ci, (h, w) in enumerate(SENT_COLS, 1):
                ws2.column_dimensions[get_column_letter(ci)].width = w
                sc(ws2, 3, ci, h, bg=C["SHDR"], fg=C["WHT"], bold=True, sz=9)

            sorted_sent = sorted(sent_results, key=lambda x: x.get("sent_score") or 0, reverse=True)
            for ri, r in enumerate(sorted_sent):
                row = ri + 4
                ws2.row_dimensions[row].height = 17
                ss  = r.get("sent_score") or 0
                sh  = _sent_hex(ss)
                sf  = "FFFFFF" if abs(ss) >= 0.2 else "1F1F1F"
                sl  = _sent_label(r.get("sent_signal"))
                bg  = C["ROW1"] if ri % 2 == 0 else C["ROW2"]
                vals = [
                    (ri + 1,                         bg,  "1F1F1F"),
                    (r.get("ticker"),                 bg,  "1F1F1F"),
                    (r.get("name"),                   bg,  "1F1F1F"),
                    (round(ss, 3),                    sh,  sf),
                    (sl,                              sh,  sf),
                    (r.get("n_articles") or 0,        bg,  "1F1F1F"),
                    (r.get("sent_momentum"),           bg,  "1F1F1F"),
                    (r.get("sent_vol_trend"),          bg,  "1F1F1F"),
                    (r.get("sent_monthly"),            bg,  "1F1F1F"),
                ]
                for ci, (val, bg_, fg_) in enumerate(vals, 1):
                    sc(ws2, row, ci, val, bg_, fg_, sz=9)

            # Headlines section
            hl_row = len(sorted_sent) + 6
            ws2.merge_cells(f"A{hl_row}:I{hl_row}")
            sc(ws2, hl_row, 1, "TOP HEADLINES", bg=C["HDR"], fg=C["WHT"], bold=True, sz=10)
            hl_row += 1

            HDR_COLS = ["#", "TICKER", "DATE", "HEADLINE", "SENTIMENT"]
            widths = [4, 9, 12, 55, 13]
            for ci, (h, w) in enumerate(zip(HDR_COLS, widths), 1):
                ws2.column_dimensions[get_column_letter(ci)].width = max(
                    ws2.column_dimensions[get_column_letter(ci)].width or 0, w)
                sc(ws2, hl_row, ci, h, bg=C["SHDR"], fg=C["WHT"], bold=True, sz=9)
            hl_row += 1

            art_n = 0
            for r in sorted_sent:
                headlines = r.get("headlines", [])[:5]
                for htext, hsentl in headlines:
                    art_n += 1
                    bg = C["ROW1"] if art_n % 2 == 0 else C["ROW2"]
                    sent_lbl = (hsentl or "").upper()
                    sh = (C["BUY"] if sent_lbl == "POSITIVE" else
                          C["SELL"] if sent_lbl == "NEGATIVE" else C["NEU"])
                    sf = "FFFFFF" if sent_lbl in ("POSITIVE", "NEGATIVE") else "1F1F1F"
                    sc(ws2, hl_row, 1, art_n,          bg,  "595959", sz=8)
                    sc(ws2, hl_row, 2, r.get("ticker"), bg,  "1F1F1F", sz=8)
                    sc(ws2, hl_row, 3, None,            bg,  "595959", sz=8)
                    sc(ws2, hl_row, 4, htext,           bg,  "1F1F1F", sz=8, h="left", wrap=True)
                    sc(ws2, hl_row, 5, sent_lbl,        sh,  sf,       sz=8)
                    ws2.row_dimensions[hl_row].height = 28
                    hl_row += 1

    # ── Buying Checklist sheet ──────────────────────────────────────────────
    has_checklist = any(r.get("checklist") for r in results)
    if has_checklist:
        ws3 = wb.create_sheet("Buying Checklist")
        ws3.sheet_view.showGridLines = False

        # Collect all checklist item names from first result that has them
        chk_names = []
        for r in results:
            if r.get("checklist"):
                chk_names = [name for name, _ in r["checklist"]]
                break

        N_CHK_LEFT  = 3   # #, Ticker, Company
        N_CHK_ITEMS = len(chk_names)
        N_CHK_RIGHT = 2   # Confidence %, Elder
        N_CHK_TOTAL = N_CHK_LEFT + N_CHK_ITEMS + N_CHK_RIGHT

        # Row 1: Title
        ws3.merge_cells(f"A1:{get_column_letter(N_CHK_TOTAL)}1")
        sc(ws3, 1, 1, "  BUYING CHECKLIST  —  Top-Down Technical Confidence",
           bg=C["HDR"], fg=C["WHT"], bold=True, sz=15)
        ws3.row_dimensions[1].height = 38

        # Row 2: Market context banner
        # Fetch market context (cached 5min by market_data module)
        try:
            from ..services.market_data import fetch_market_context
            mkt_ctx = fetch_market_context()
        except Exception:
            mkt_ctx = {}
        vix_val  = mkt_ctx.get("vix")
        vix_safe = mkt_ctx.get("vix_safe")
        nl_val   = mkt_ctx.get("nyse_new_lows")
        nl_safe  = mkt_ctx.get("breadth_safe")

        ctx_parts = []
        if vix_val is not None:
            s = "✓" if vix_safe else "✗"
            ctx_parts.append(f"VIX: {vix_val:.1f} ({s} {'Safe' if vix_safe else 'Elevated'})")
        if nl_val is not None:
            s = "✓" if nl_safe else "✗"
            ctx_parts.append(f"NYSE New Lows: {nl_val} ({s} {'< 500' if nl_safe else '≥ 500'})")
        ctx_str = "  │  ".join(ctx_parts) if ctx_parts else "Market context unavailable"

        ws3.merge_cells(f"A2:{get_column_letter(N_CHK_TOTAL)}2")
        mctx_bg = C["BUY"] if (vix_safe and nl_safe) else C["SELL"] if (vix_safe is False or nl_safe is False) else C["SHDR"]
        sc(ws3, 2, 1, f"  MARKET CONTEXT  │  {ctx_str}",
           bg=mctx_bg, fg=C["WHT"], sz=10, italic=True)
        ws3.row_dimensions[2].height = 24

        # Row 3: Separator
        ws3.row_dimensions[3].height = 6
        for ci in range(1, N_CHK_TOTAL + 1):
            ws3.cell(row=3, column=ci).fill = fill(C["HDR"])

        # Row 4: Column headers
        ws3.row_dimensions[4].height = 28
        left_hdrs = [("#", 4), ("TICKER", 9), ("COMPANY", 20)]
        for ci, (h, w) in enumerate(left_hdrs, 1):
            ws3.column_dimensions[get_column_letter(ci)].width = w
            sc(ws3, 4, ci, h, bg=C["SHDR"], fg=C["WHT"], bold=True, sz=9)

        for ci, name in enumerate(chk_names, N_CHK_LEFT + 1):
            ws3.column_dimensions[get_column_letter(ci)].width = 11
            sc(ws3, 4, ci, name, bg=C["SEC"], fg=C["WHT"], bold=True, sz=8, wrap=True)

        # Confidence + Elder headers
        conf_ci = N_CHK_LEFT + N_CHK_ITEMS + 1
        elder_ci = conf_ci + 1
        ws3.column_dimensions[get_column_letter(conf_ci)].width = 12
        ws3.column_dimensions[get_column_letter(elder_ci)].width = 10
        sc(ws3, 4, conf_ci, "CONFIDENCE", bg=C["HDR"], fg=C["WHT"], bold=True, sz=9)
        sc(ws3, 4, elder_ci, "ELDER", bg=C["HDR"], fg=C["WHT"], bold=True, sz=9)

        # Data rows
        for ri, r in enumerate(results):
            row = 5 + ri
            ws3.row_dimensions[row].height = 20
            bg = C["ROW2"] if ri % 2 == 0 else C["ROW1"]

            sc(ws3, row, 1, ri + 1, bg, "1F1F1F", sz=9)
            sc(ws3, row, 2, r.get("ticker"), bg, "1F1F1F", bold=True, sz=9)
            sc(ws3, row, 3, r.get("name"), bg, "1F1F1F", sz=9, h="left")

            checks = r.get("checklist", [])
            for ci2, (_, passed) in enumerate(checks, N_CHK_LEFT + 1):
                if passed is None:
                    sc(ws3, row, ci2, "—", bg, C["GRY"], sz=9)
                elif passed:
                    sc(ws3, row, ci2, "✓", C["BUY"], C["WHT"], bold=True, sz=11)
                else:
                    sc(ws3, row, ci2, "✗", C["SELL"], C["WHT"], bold=True, sz=11)

            # Confidence %
            conf = r.get("confidence")
            if conf is not None:
                conf_bg = (C["SBUY"] if conf >= 80 else C["BUY"] if conf >= 60
                           else C["NEU"] if conf >= 40 else C["SELL"] if conf >= 20
                           else C["SSELL"])
                conf_fg = C["WHT"] if conf >= 60 or conf < 20 else C["BLK"]
                sc(ws3, row, conf_ci, f"{conf:.0f}%", conf_bg, conf_fg, bold=True, sz=11)
            else:
                sc(ws3, row, conf_ci, "N/A", bg, C["GRY"], sz=9)

            # Elder Impulse (combined daily/weekly)
            elder_d = r.get("elder_d", "blue")
            elder_w = r.get("elder_w", "blue")
            ELDER_BG = {"green": C["BUY"], "red": C["SELL"], "blue": C["SHDR"]}
            ELDER_LBL = {"green": "▲", "red": "▼", "blue": "●"}
            lbl = f"W{ELDER_LBL.get(elder_w, '●')} D{ELDER_LBL.get(elder_d, '●')}"
            # Use worst of the two for background
            if elder_d == "red" or elder_w == "red":
                ebg = C["SELL"]
            elif elder_d == "green" and elder_w == "green":
                ebg = C["BUY"]
            else:
                ebg = C["SHDR"]
            sc(ws3, row, elder_ci, lbl, ebg, C["WHT"], bold=True, sz=10)

        # Footer
        fr = 5 + len(results)
        ws3.row_dimensions[fr].height = 14
        ws3.merge_cells(f"A{fr}:{get_column_letter(N_CHK_TOTAL)}{fr}")
        fc = ws3.cell(row=fr, column=1)
        fc.value = (
            "✓ = condition met  │  ✗ = condition not met  │  "
            "Confidence = % of conditions passed  │  "
            "Elder: W = weekly, D = daily (▲ green = bullish, ▼ red = bearish, ● blue = neutral)"
        )
        fc.fill = fill(C["HDR"])
        fc.font = font("AAAAAA", size=8)
        fc.alignment = align("center", "center")

    # ── Individual stock sheets ───────────────────────────────────────────────
    for r in results:
        ticker = r.get("ticker", "")
        sname  = ticker[:31]
        ws_s   = wb.create_sheet(sname)

        sym = r.get("symbol", "$")
        px  = r.get("price")
        css = r.get("signal_css", "sig-neu")
        sig_bg  = _sig_hex(css)
        sig_fg  = _sig_txt(css)
        sig_lbl = _sig_label(r.get("signal") or "")

        ws_s.column_dimensions["A"].width = 12
        ws_s.column_dimensions["B"].width = 28
        ws_s.column_dimensions["C"].width = 12
        ws_s.column_dimensions["D"].width = 12

        # Title
        ws_s.row_dimensions[1].height = 28
        ws_s.merge_cells("A1:D1")
        sc(ws_s, 1, 1, f"{r.get('name', ticker)}  ·  {r.get('sector', '')}",
           bg=C["HDR"], fg=C["WHT"], bold=True, sz=13)

        ws_s.row_dimensions[2].height = 16
        ws_s.merge_cells("A2:D2")
        sc(ws_s, 2, 1, f"{r.get('ticker')}  ·  {r.get('name', '')}  ·  {r.get('currency', 'USD')}",
           bg=C["SHDR"], fg=C["WHT"], sz=9, italic=True)

        # Quick-stat boxes: row 3 headers, row 4 values
        def stat_box(row_h, row_v, col, label, value, bg_h=C["SHDR"], bg_v=C["ROW2"],
                     bold_v=False, fg_v="1F1F1F"):
            sc(ws_s, row_h, col, label, bg_h, C["WHT"], bold=True, sz=9)
            sc(ws_s, row_v, col, value, bg_v, fg_v, bold_v, sz=11)

        ws_s.row_dimensions[3].height = 22
        ws_s.row_dimensions[4].height = 26

        # Use contextual signal for per-stock display
        _ctx_sig = r.get("ctx_signal") or sig_lbl
        if "AVOID" in _ctx_sig:
            _s_bg, _s_fg = _sig_hex("ssell"), _sig_txt("ssell")
        elif _ctx_sig.startswith("HOLD"):
            _s_bg, _s_fg = _sig_hex("neu"), _sig_txt("neu")
        else:
            _s_bg, _s_fg = sig_bg, sig_fg

        px_str = f"{sym}{round(px, 2)}" if px else "N/A"
        stat_box(3, 4, 1, "PRICE", px_str, bold_v=True)
        stat_box(3, 4, 2, "SIGNAL", _ctx_sig, bg_v=_s_bg, fg_v=_s_fg, bold_v=True)
        raw = r.get("raw_score")
        stat_box(3, 4, 3, "TECH SCORE", f"{raw:+.0f}" if raw is not None else "N/A", bold_v=True)
        stat_box(3, 4, 4, "OVERALL", f"{(r.get('overall_score') or 0):.3f}", bold_v=True)

        # Second quick-stat row: returns
        ws_s.row_dimensions[5].height = 22
        ws_s.row_dimensions[6].height = 22

        def _fmt_ret(v):
            if v is None: return "N/A"
            return f"+{v:.1f}%" if v > 0 else f"{v:.1f}%"

        stat_box(5, 6, 1, "1D",  _fmt_ret(r.get("ret_1d")))
        stat_box(5, 6, 2, "1W",  _fmt_ret(r.get("ret_1w")))
        stat_box(5, 6, 3, "1M",  _fmt_ret(r.get("ret_1m")))
        stat_box(5, 6, 4, "3M",  _fmt_ret(r.get("ret_3m")))

        # Third row: technical
        ws_s.row_dimensions[7].height = 22
        ws_s.row_dimensions[8].height = 22

        cross = r.get("ma_cross")
        cross_lbl = "GOLDEN ✓" if cross == "golden" else "DEATH ✗" if cross == "death" else "N/A"
        rsi = r.get("rsi")
        stat_box(7, 8, 1, "RSI 14",   f"{rsi:.1f}" if rsi else "N/A")
        stat_box(7, 8, 2, "MA CROSS", cross_lbl)
        bb = r.get("bb_pct")
        stat_box(7, 8, 3, "BOLL %",   f"{bb:.1f}%" if bb else "N/A")
        stat_box(7, 8, 4, "52W POS",  f"{r.get('w52_pct', 0):.1f}%" if r.get("w52_pct") is not None else "N/A")

        # Fourth row: fundamentals
        ws_s.row_dimensions[9].height = 22
        ws_s.row_dimensions[10].height = 22

        pe = r.get("pe_trail")
        pef = r.get("pe_fwd")
        peg = r.get("peg")
        stat_box(9, 10, 1, "TRAIL P/E", f"{pe:.1f}" if pe else "N/A")
        stat_box(9, 10, 2, "FWD P/E",   f"{pef:.1f}" if pef else "N/A")
        stat_box(9, 10, 3, "PEG",        f"{peg:.2f}" if peg else "N/A")
        stat_box(9, 10, 4, "NET MARGIN", f"{r.get('net_mgn'):+.1f}%" if r.get("net_mgn") is not None else "N/A")

        # Fifth row: Regime + Quant indicators
        ws_s.row_dimensions[11].height = 22
        ws_s.row_dimensions[12].height = 22

        _adx_v = r.get("adx")
        _regime_v = r.get("regime", "NEUTRAL")
        _regime_colors = {"TREND": C["SBUY"], "MEAN_REVERSION": C["SHDR"], "NEUTRAL": C["NEU"]}
        stat_box(11, 12, 1, "REGIME", _regime_v,
                 bg_v=_regime_colors.get(_regime_v, C["GRY"]),
                 fg_v=C["WHT"] if _regime_v != "NEUTRAL" else C["BLK"], bold_v=True)
        stat_box(11, 12, 2, "ADX", f"{_adx_v:.1f}" if _adx_v else "N/A", bold_v=True)
        _atr_pct_v = r.get("atr_pct")
        stat_box(11, 12, 3, "ATR %", f"{_atr_pct_v:.1f}%" if _atr_pct_v else "N/A")
        _rs_v = r.get("rs_1m")
        stat_box(11, 12, 4, "RS vs SPY", f"{_rs_v:+.1f}%" if _rs_v is not None else "N/A",
                 bg_v=C["BUY"] if _rs_v and _rs_v > 2 else C["SELL"] if _rs_v and _rs_v < -2 else C["ROW2"],
                 fg_v=C["WHT"] if _rs_v and abs(_rs_v) > 2 else C["BLK"], bold_v=True)

        # Sixth row: Trend stage + Vol regime + Mkt regime + Regime change
        ws_s.row_dimensions[13].height = 22
        ws_s.row_dimensions[14].height = 22

        _ts_v = r.get("trend_stage")
        _ts_colors = {"EARLY": C["BUY"], "HEALTHY": C["SBUY"], "EXTENDED": C["NEU"],
                       "OVEREXTENDED": C["SELL"], "PARABOLIC": C["SSELL"]}
        stat_box(13, 14, 1, "TREND STAGE", _ts_v or "N/A",
                 bg_v=_ts_colors.get(_ts_v, C["GRY"]),
                 fg_v=C["WHT"] if _ts_v in ("EARLY", "HEALTHY", "OVEREXTENDED", "PARABOLIC") else C["BLK"],
                 bold_v=True)
        _vr_v = r.get("vol_regime")
        _vr_colors = {"LOW": C["BUY"], "NORMAL": C["NEU"], "HIGH": C["SELL"], "EXTREME": C["SSELL"]}
        stat_box(13, 14, 2, "VOL REGIME", _vr_v or "N/A",
                 bg_v=_vr_colors.get(_vr_v, C["GRY"]),
                 fg_v=C["WHT"] if _vr_v in ("LOW", "HIGH", "EXTREME") else C["BLK"],
                 bold_v=True)
        _mr_v = r.get("mkt_regime")
        _mr_colors = {"BULLISH": C["SBUY"], "BEARISH": C["SSELL"], "TRANSITION": C["NEU"]}
        stat_box(13, 14, 3, "MKT REGIME", _mr_v or "N/A",
                 bg_v=_mr_colors.get(_mr_v, C["GRY"]),
                 fg_v=C["WHT"] if _mr_v in ("BULLISH", "BEARISH") else C["BLK"],
                 bold_v=True)
        _rc_v = r.get("regime_chg")
        stat_box(13, 14, 4, "REGIME CHG", _rc_v or "—",
                 bg_v=C["NEU"], fg_v=C["BLK"])

        # Seventh row: Momentum + Risk (v4)
        ws_s.row_dimensions[15].height = 22
        ws_s.row_dimensions[16].height = 22

        _mom_v = r.get("momentum_score", 0)
        _mom_bg = C["SBUY"] if _mom_v >= 0.65 else C["BUY"] if _mom_v >= 0.40 else C["NEU"]
        stat_box(15, 16, 1, "MOMENTUM", f"{_mom_v:.2f}",
                 bg_v=_mom_bg, fg_v=C["WHT"] if _mom_v >= 0.40 else C["BLK"], bold_v=True)
        _risk_v = r.get("risk_score", 0)
        _risk_bg = C["SSELL"] if _risk_v >= 1.0 else C["SELL"] if _risk_v >= 0.5 else C["NEU"] if _risk_v >= 0.15 else C["BUY"]
        stat_box(15, 16, 2, "RISK", f"{_risk_v:.2f}",
                 bg_v=_risk_bg, fg_v=C["WHT"] if _risk_v >= 0.5 else C["BLK"], bold_v=True)
        # Context hint
        _ctx_hint = r.get("ctx_hint", "")
        stat_box(15, 16, 3, "CONTEXT", _ctx_hint or "—")
        stat_box(15, 16, 4, "CTX SIGNAL", r.get("ctx_signal") or sig_lbl,
                 bg_v=_s_bg, fg_v=_s_fg, bold_v=True)

        # Eighth row: Elder Impulse + Confidence
        ws_s.row_dimensions[17].height = 22
        ws_s.row_dimensions[18].height = 22

        _elder_bg = {"green": C["BUY"], "red": C["SELL"], "blue": C["SHDR"]}
        _elder_lbl = {"green": "▲ BULL", "red": "▼ BEAR", "blue": "● NEUTRAL"}
        e_d = r.get("elder_d", "blue")
        e_w = r.get("elder_w", "blue")
        stat_box(17, 18, 1, "ELDER DAILY",  _elder_lbl.get(e_d, "N/A"),
                 bg_v=_elder_bg.get(e_d, C["SHDR"]), fg_v=C["WHT"], bold_v=True)
        stat_box(17, 18, 2, "ELDER WEEKLY", _elder_lbl.get(e_w, "N/A"),
                 bg_v=_elder_bg.get(e_w, C["SHDR"]), fg_v=C["WHT"], bold_v=True)
        conf = r.get("adj_confidence") or r.get("confidence")
        conf_bg = (C["SBUY"] if conf and conf >= 80 else C["BUY"] if conf and conf >= 60
                   else C["NEU"] if conf and conf >= 40 else C["SELL"] if conf and conf is not None
                   else C["GRY"])
        stat_box(17, 18, 3, "ADJ CONF", f"{conf:.0f}%" if conf is not None else "N/A",
                 bg_v=conf_bg, fg_v=C["WHT"] if conf and conf >= 60 else C["BLK"], bold_v=True)
        chk_p, chk_t = r.get("chk_passed", 0), r.get("chk_total", 0)
        stat_box(17, 18, 4, "CHECKLIST", f"{chk_p}/{chk_t} passed")

        # Ninth row: sentiment (if available)
        ws_s.row_dimensions[19].height = 22
        ws_s.row_dimensions[20].height = 22

        ss = r.get("sent_score")
        sl = _sent_label(r.get("sent_signal"))
        sh = _sent_hex(ss)
        stat_box(19, 20, 1, "SENT SCORE",  f"{ss:.3f}" if ss is not None else "N/A",
                 bg_v=sh, fg_v="FFFFFF" if ss is not None and abs(ss) >= 0.2 else "1F1F1F")
        stat_box(19, 20, 2, "SENT SIGNAL", sl,
                 bg_v=sh, fg_v="FFFFFF" if ss is not None and abs(ss) >= 0.2 else "1F1F1F")
        stat_box(19, 20, 3, "ARTICLES", str(r.get("n_articles") or 0))
        stat_box(19, 20, 4, "MOMENTUM", f"{r.get('sent_momentum', 0):.3f}" if r.get("sent_momentum") is not None else "N/A")

        ws_s.freeze_panes = "A22"

        # ── Write historical data from DataFrame ──────────────────────────────
        DATA_HDRS = [
            "DATE", "CLOSE", "MA20", "MA50", "MA200", "RSI",
            "MACD", "BOLL %", "VOL RATIO", "ATR", "RET 1D%",
        ]

        ws_s.row_dimensions[21].height = 20
        for ci, h in enumerate(DATA_HDRS, 1):
            if ci > 4:
                ws_s.column_dimensions[get_column_letter(ci)].width = 10
            sc(ws_s, 21, ci, h, bg=C["SHDR"], fg=C["WHT"], bold=True, sz=8)

        # Retrieve DataFrame from raw_data
        df = None
        if raw_data and ticker in raw_data:
            rd = raw_data[ticker]
            if rd and hasattr(rd, 'get') and "df" in rd:
                df = rd["df"]

        n_hist_rows = 0
        if df is not None and len(df) > 0:
            df_slice = df.tail(252).copy()
            n_hist_rows = len(df_slice)

            for ri2, (idx, row_data) in enumerate(df_slice.iterrows(), 22):
                bg = C["ROW1"] if ri2 % 2 == 0 else C["ROW2"]
                ws_s.row_dimensions[ri2].height = 14
                date_str = idx.strftime("%Y-%m-%d") if hasattr(idx, "strftime") else str(idx)

                def _sv(col_name):
                    v = row_data.get(col_name)
                    if v is None: return None
                    try:
                        f = float(v)
                        return None if (np.isnan(f) or np.isinf(f)) else round(f, 4)
                    except (TypeError, ValueError):
                        return None

                cells_visible = [
                    (date_str,          bg, "595959", None),
                    (_sv("Close"),      bg, "1F1F1F", f"{sym}#,##0.00"),
                    (_sv("MA20"),       bg, "1F1F1F", "0.00"),
                    (_sv("MA50"),       bg, "1F1F1F", "0.00"),
                    (_sv("MA200"),      bg, "1F1F1F", "0.00"),
                    (_sv("RSI"),        bg, "1F1F1F", "0.0"),
                    (_sv("MACD"),       bg, "1F1F1F", "0.0000"),
                    (_sv("BB_Pct"),     bg, "1F1F1F", "0.0"),
                    (_sv("Vol_Ratio"),  bg, "1F1F1F", "0.00"),
                    (_sv("ATR"),        bg, "1F1F1F", "0.00"),
                    (_sv("Ret_1D"),     bg, "1F1F1F", '0.00"%"'),
                ]
                for ci, (val, bg_, fg_, nf) in enumerate(cells_visible, 1):
                    sc(ws_s, ri2, ci, val, bg_, fg_, sz=8, nfmt=nf)

        data_end_row = 21 + n_hist_rows

        # ── Render matplotlib chart and embed as image ────────────────────────
        if df is not None and n_hist_rows > 20:
            try:
                chart_bytes = _render_stock_chart(df.tail(252).copy(), ticker, sym)
                if chart_bytes:
                    from openpyxl.drawing.image import Image as XlImage
                    img = XlImage(io.BytesIO(chart_bytes))
                    # Place chart to the right of stat boxes (column F, row 1)
                    ws_s.add_image(img, "F1")
            except Exception:
                pass  # chart generation failed silently; data table still present

        # ── News articles section ─────────────────────────────────────────────
        articles = r.get("articles", [])
        if articles:
            art_start = max(21, data_end_row + 2) if n_hist_rows > 0 else 21
            ws_s.row_dimensions[art_start].height = 22
            ws_s.merge_cells(f"A{art_start}:K{art_start}")
            sc(ws_s, art_start, 1, "NEWS ARTICLES", bg=C["HDR"], fg=C["WHT"], bold=True, sz=11)

            art_start += 1
            ws_s.merge_cells(f"A{art_start}:K{art_start}")
            na = r.get("n_articles", 0)
            np_ = r.get("n_positive", 0)
            nn = r.get("n_negative", 0)
            sc(ws_s, art_start, 1,
               f"{na} articles  ·  {np_} positive  ·  {nn} negative  ·  {na - np_ - nn} neutral",
               bg=C["SHDR"], fg=C["WHT"], sz=9, italic=True)

            art_start += 1
            ART_HDRS = ["#", "DATE", "HEADLINE", "SOURCE", "SENTIMENT"]
            art_widths = [4, 12, 60, 16, 12]
            for ci, (h, w) in enumerate(zip(ART_HDRS, art_widths), 1):
                ws_s.column_dimensions[get_column_letter(ci)].width = max(
                    ws_s.column_dimensions[get_column_letter(ci)].width or 0, w)
                sc(ws_s, art_start, ci, h, bg=C["SHDR"], fg=C["WHT"], bold=True, sz=9)

            art_start += 1
            sorted_arts = sorted(articles, key=lambda a: a.get("date", ""), reverse=True)
            for ai, art in enumerate(sorted_arts):
                bg = C["ROW1"] if ai % 2 == 0 else C["ROW2"]
                slbl = (art.get("sentiment") or "").upper()
                sh = (C["BUY"] if slbl == "POSITIVE" else
                      C["SELL"] if slbl == "NEGATIVE" else C["NEU"])
                sf = "FFFFFF" if slbl in ("POSITIVE", "NEGATIVE") else "1F1F1F"
                sc(ws_s, art_start, 1, ai + 1,               bg,  "595959", sz=8)
                sc(ws_s, art_start, 2, art.get("date", ""),   bg,  "595959", sz=8)
                sc(ws_s, art_start, 3, art.get("title", ""),  bg,  "1F1F1F", sz=8, h="left", wrap=True)
                sc(ws_s, art_start, 4, art.get("source", ""), bg,  "595959", sz=8)
                sc(ws_s, art_start, 5, slbl,                  sh,  sf,       sz=8)
                ws_s.row_dimensions[art_start].height = 24
                art_start += 1

    # ── Save to bytes ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Route ─────────────────────────────────────────────────────────────────────

@bp.route("/excel", methods=["POST"])
def excel():
    body = request.get_json(force=True, silent=True) or {}
    results = body.get("results")
    config  = body.get("config", {})
    task_id = body.get("task_id")
    save_path = body.get("save_path")  # optional: full path to save file

    if not results:
        return jsonify({"error": "No results provided"}), 400

    # Retrieve raw DataFrames for chart generation
    raw_data = None
    if task_id:
        from .analysis import _tasks, _tasks_lock
        with _tasks_lock:
            task = _tasks.get(task_id, {})
            raw_data = task.get("_raw_data", {})

    try:
        xlsx_bytes = _build_excel(results, config, raw_data=raw_data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    # If a save_path was provided, write directly to disk
    if save_path:
        import os
        try:
            os.makedirs(os.path.dirname(save_path), exist_ok=True)
            with open(save_path, "wb") as f:
                f.write(xlsx_bytes)
            return jsonify({"ok": True, "path": save_path})
        except Exception as e:
            return jsonify({"error": f"Failed to save: {e}"}), 500

    fname = f"stock_analysis_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return Response(
        xlsx_bytes,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
